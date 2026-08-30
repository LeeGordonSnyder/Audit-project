"""
Excel output. Two entry points:

- build_demo_template(path): generates a workbook that mimics the layout of
  the example deployment sheet (header, break columns, hourly grid, role
  legend, a placeholder Hourly Tracker section). Useful for trying the tool
  end-to-end without a real workbook, and as a starting point to compare
  against your real file's exact cell layout.

- fill_workbook(template_path, output_path, shifts, ...): opens a workbook
  (your real deployment master, or the demo template), writes the roster,
  break times, and colored hourly grid into the region described in
  config.py, and leaves every other cell (goals, Hourly Tracker formulas,
  legend, etc.) exactly as it was.
"""
from __future__ import annotations

from datetime import date as date_cls
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import config
from .breaks import Shift

THIN = Side(style="thin", color="BFBFBF")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")


def format_time(minutes: int) -> str:
    h24, m = divmod(minutes, 60)
    period = "am" if h24 < 12 else "pm"
    h12 = h24 % 12
    if h12 == 0:
        h12 = 12
    return f"{h12}{period}" if m == 0 else f"{h12}:{m:02d}{period}"


def format_shift_range(shift: Shift) -> str:
    return f"{format_time(shift.start_min)} - {format_time(shift.end_min)}"


def _role_style(code: str):
    style = config.ROLE_STYLES.get(code)
    if style is None:
        raise ValueError(f"Unknown role code {code!r}; is it in config.ROLE_STYLES?")
    return style


def _apply_role_cell(cell, code: str, text: str | None = None):
    style = _role_style(code)
    cell.value = text if text is not None else style.code
    cell.fill = PatternFill(start_color=style.fill_hex, end_color=style.fill_hex, fill_type="solid")
    cell.font = Font(color=style.font_hex, bold=True, size=9)
    cell.alignment = CENTER
    cell.border = CELL_BORDER


def hour_overlaps_shift(hour: int, shift: Shift) -> bool:
    hour_start = hour * 60
    hour_end = hour_start + 60
    return shift.start_min < hour_end and shift.end_min > hour_start


def hour_overlaps_break(hour: int, shift: Shift) -> bool:
    hour_start = hour * 60
    hour_end = hour_start + 60
    return any(hour_start < b.end_min and (b.start_min + b.duration_min) > hour_start and
               b.start_min < hour_end for b in shift.breaks)


def fill_workbook(
    template_path: str | Path,
    output_path: str | Path,
    shifts: list[Shift],
    sheet_name: str | None = None,
    target_date: date_cls | None = None,
    date_cell: str | None = None,
    show_breaks_in_grid: bool = True,
) -> list[str]:
    """Fills the roster/break/hourly-grid region of a workbook with the given
    shifts. Returns a list of warning strings (e.g. flagged break placements,
    shifts that didn't fit in the reserved rows) for the caller to surface.
    """
    warnings: list[str] = []
    wb = openpyxl.load_workbook(template_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    if date_cell and target_date is not None:
        ws[date_cell] = f"{target_date:%A, %b} {target_date.day}"

    ordered = sorted(shifts, key=lambda s: s.start_min)
    if len(ordered) > config.STAFF_ROW_CAPACITY:
        warnings.append(
            f"{len(ordered)} shifts but only {config.STAFF_ROW_CAPACITY} rows reserved; "
            f"extra shifts were dropped. Raise config.STAFF_ROW_CAPACITY."
        )
        ordered = ordered[: config.STAFF_ROW_CAPACITY]

    for i, shift in enumerate(ordered):
        row = config.STAFF_START_ROW + i

        name_cell = ws.cell(row=row, column=config.COL_NAME, value=shift.name)
        name_cell.border = CELL_BORDER
        shift_cell = ws.cell(row=row, column=config.COL_SHIFT, value=format_shift_range(shift))
        shift_cell.border = CELL_BORDER
        shift_cell.alignment = CENTER

        fifteens = [b for b in shift.breaks if b.duration_min == 15]
        thirty = next((b for b in shift.breaks if b.duration_min == 30), None)

        break_cols = [
            (config.COL_BREAK_30, thirty),
            (config.COL_BREAK_15_A, fifteens[0] if len(fifteens) > 0 else None),
            (config.COL_BREAK_15_B, fifteens[1] if len(fifteens) > 1 else None),
        ]
        for col, brk in break_cols:
            cell = ws.cell(row=row, column=col)
            cell.border = CELL_BORDER
            cell.alignment = CENTER
            if brk is None:
                continue
            cell.value = format_time(brk.start_min)
            if brk.flagged:
                cell.font = Font(color="CC0000", bold=True)
                warnings.append(
                    f"{shift.name}: could not find a fully valid slot for their "
                    f"{brk.duration_min}-min break; placed at {format_time(brk.start_min)} "
                    f"-- please review floor coverage manually."
                )

        for hour in range(config.GRID_START_HOUR, config.GRID_END_HOUR + 1):
            col = config.GRID_START_COL + (hour - config.GRID_START_HOUR)
            cell = ws.cell(row=row, column=col)
            cell.border = CELL_BORDER
            if show_breaks_in_grid and hour_overlaps_break(hour, shift):
                _apply_role_cell(cell, "B")
            elif hour_overlaps_shift(hour, shift):
                _apply_role_cell(cell, shift.role)

    wb.save(output_path)
    return warnings


def build_demo_template(path: str | Path) -> None:
    """Generates a workbook that mimics the example deployment sheet's
    layout, for demoing/testing the tool without a real workbook on hand."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Demo"

    ws.column_dimensions[get_column_letter(config.COL_NAME)].width = 16
    ws.column_dimensions[get_column_letter(config.COL_SHIFT)].width = 18
    for col in (config.COL_BREAK_30, config.COL_BREAK_15_A, config.COL_BREAK_15_B):
        ws.column_dimensions[get_column_letter(col)].width = 9
    for hour in range(config.GRID_START_HOUR, config.GRID_END_HOUR + 1):
        col = config.GRID_START_COL + (hour - config.GRID_START_HOUR)
        ws.column_dimensions[get_column_letter(col)].width = 7

    header_fill = PatternFill(start_color="9E3A5C", end_color="9E3A5C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.merge_cells(start_row=1, start_column=config.COL_NAME, end_row=1, end_column=config.COL_SHIFT)
    top = ws.cell(row=1, column=config.COL_NAME, value="DATE")
    top.fill, top.font, top.alignment = header_fill, header_font, CENTER

    ws.merge_cells(start_row=1, start_column=config.COL_BREAK_30, end_row=1, end_column=config.COL_BREAK_15_B)
    date_cell = ws.cell(row=1, column=config.COL_BREAK_30, value="(fill in date)")
    date_cell.alignment = CENTER

    last_grid_col = config.GRID_START_COL + config.GRID_NUM_COLS - 1
    ws.merge_cells(start_row=1, start_column=config.GRID_START_COL, end_row=1, end_column=last_grid_col)
    goal_cell = ws.cell(row=1, column=config.GRID_START_COL, value="Goal / Week / Month figures go here (untouched by the sync script)")
    goal_cell.fill, goal_cell.font, goal_cell.alignment = header_fill, header_font, CENTER

    headers_row3 = {
        config.COL_NAME: "TEAM",
        config.COL_SHIFT: "Shift",
        config.COL_BREAK_30: "BREAK",
        config.COL_BREAK_15_A: "15 MIN",
        config.COL_BREAK_15_B: "15 MIN",
    }
    for col, label in headers_row3.items():
        c = ws.cell(row=config.HOUR_HEADER_ROW, column=col, value=label)
        c.font = Font(bold=True)
        c.alignment = CENTER
        c.border = CELL_BORDER
    for hour in range(config.GRID_START_HOUR, config.GRID_END_HOUR + 1):
        col = config.GRID_START_COL + (hour - config.GRID_START_HOUR)
        c = ws.cell(row=config.HOUR_HEADER_ROW, column=col, value=format_time(hour * 60))
        c.font = Font(bold=True)
        c.alignment = CENTER
        c.border = CELL_BORDER

    for r in range(config.STAFF_START_ROW, config.STAFF_START_ROW + config.STAFF_ROW_CAPACITY):
        for col in list(headers_row3) + list(
            range(config.GRID_START_COL, config.GRID_START_COL + config.GRID_NUM_COLS)
        ):
            ws.cell(row=r, column=col).border = CELL_BORDER

    tracker_row = config.STAFF_START_ROW + config.STAFF_ROW_CAPACITY + 2
    ws.merge_cells(
        start_row=tracker_row, start_column=config.COL_NAME,
        end_row=tracker_row, end_column=last_grid_col,
    )
    tc = ws.cell(row=tracker_row, column=config.COL_NAME, value="Hourly Tracker (sales/goal figures -- untouched by the sync script)")
    tc.fill, tc.font, tc.alignment = header_fill, header_font, CENTER

    legend_col = last_grid_col + 2
    lc = ws.cell(row=1, column=legend_col, value="Shift Functions")
    lc.font = Font(bold=True)
    for i, style in enumerate(config.ROLE_STYLES.values(), start=2):
        code_cell = ws.cell(row=i, column=legend_col, value=style.code)
        _apply_role_cell(code_cell, style.code)
        ws.cell(row=i, column=legend_col + 1, value=style.label)
    ws.column_dimensions[get_column_letter(legend_col + 1)].width = 22

    wb.save(path)
